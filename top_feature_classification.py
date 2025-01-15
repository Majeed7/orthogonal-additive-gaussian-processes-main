import os
import pickle
import numpy as np
import pandas as pd
from sklearn.svm import SVC, SVR
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.metrics import accuracy_score, roc_auc_score, precision_score, recall_score, f1_score
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score, explained_variance_score
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook, Workbook
from sklearn.preprocessing import LabelEncoder
from sklearn.utils.multiclass import type_of_target
import sys

from real_datasets import load_dataset

def reload_dataset(dataset_name):
    """Reload the dataset by its name."""
    # Assuming you have a similar function like `load_dataset` in your earlier code
    return load_dataset(dataset_name)


def calculate_classification_scores(y_true, y_pred, y_prob):
    """Calculate 5 classification metrics."""
    return {
        "Accuracy": accuracy_score(y_true, y_pred),
        "AUROC": roc_auc_score(y_true, y_prob) if y_prob is not None else None,
        "Precision": precision_score(y_true, y_pred),
        "Recall": recall_score(y_true, y_pred),
        "F1-Score": f1_score(y_true, y_pred)
    }


def calculate_regression_scores(y_true, y_pred):
    """Calculate 5 regression metrics."""
    return {
        "MSE": mean_squared_error(y_true, y_pred),
        "MAE": mean_absolute_error(y_true, y_pred),
        "R-Squared": r2_score(y_true, y_pred),
        "RMSE": mean_squared_error(y_true, y_pred),
        "Explained Variance": explained_variance_score(y_true, y_pred)
    }


from sklearn.model_selection import GridSearchCV

def train_svm_on_selected_features(X_train, y_train, X_test, y_test, is_classification):
    """
    Train an SVM model with hyperparameter tuning and evaluate its performance.

    Parameters:
        X_train: Training feature matrix
        y_train: Training labels
        X_test: Testing feature matrix
        y_test: Testing labels
        is_classification: Boolean indicating if the task is classification

    Returns:
        dict: Dictionary containing performance scores
    """
    # Define the parameter grid for fine-tuning
    if is_classification:
        param_grid = {
            "svm__C": [0.1, 1, 10, 100],  # Regularization parameter
            #"svm__gamma": ["scale", "auto"],  # Kernel coefficient
         }
    else:
        param_grid = {
            "svm__C": [0.1, 1, 10, 100],  # Regularization parameter
            #"svm__gamma": ["scale", "auto"],  # Kernel coefficient
            'svm__epsilon': [0.1, 0.2, 0.5]
        }

    # Create the SVM model
    model = SVC(probability=True) if is_classification else SVR()

    # Create the pipeline
    pipeline = Pipeline([
        ("imputer", SimpleImputer(strategy="mean")),  # Handle missing values
        ("svm", model)  # SVM model
    ])

    # Perform Grid Search for hyperparameter tuning
    grid_search = GridSearchCV(pipeline, param_grid, cv=3, scoring="accuracy" if is_classification else "neg_mean_squared_error")
    grid_search.fit(X_train, y_train)

    # Get the best estimator
    best_model = grid_search.best_estimator_

    # Predict on the test set
    y_pred = best_model.predict(X_test)
    y_prob = best_model.predict_proba(X_test)[:, 1] if is_classification else None

    # Evaluate performance
    if is_classification:
        scores = calculate_classification_scores(y_test, y_pred, y_prob)
    else:
        scores = calculate_regression_scores(y_test, y_pred)

    return scores


def main():

    # Check if an argument is passed
    if len(sys.argv) < 2:
        print("No argument passed. Using default value: 0.1")
        top_precent = 0.1  # Set a default value if no argument is provided
    else:
        # Retrieve the parameter passed from Bash
        parameter = sys.argv[1]

        # Try converting the argument to a number
        try:
            # Try converting to an integer
            top_precent = float(parameter)
        except ValueError:
            # If it fails, try converting to a float
            top_precent = 0.1
            print("Cannot process the value. Using default value: 0.1")

    # Load the Excel file with feature importance data
    feature_importance_file = "results/real_feature_importances.xlsx"
    wb = load_workbook(feature_importance_file)

    # Create a new workbook for storing results
    results_wb = Workbook()

    for sheet_name in wb.sheetnames:
        print(f"Processing dataset: {sheet_name}")
        if sheet_name in ['keggdirected']: continue
        sheet = wb[sheet_name]

        # Reload the dataset
        X, y = reload_dataset(sheet_name)

        # Determine if it's a classification or regression task
        is_classification = type_of_target(y) in ["binary", "multiclass"]
        if is_classification:
            y = LabelEncoder().fit_transform(y)

        # Train-test split
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1, random_state=42)

        result_sheet = results_wb.create_sheet(title=sheet_name)

        # Set column titles dynamically
        score_titles = ["Accuracy", "AUROC", "Precision", "Recall", "F1-Score"] if is_classification else \
                       ["MSE", "MAE", "R-Squared", "RMSE", "Explained Variance"]
        result_sheet.append(["Feature Selector"] + score_titles)

        # Process each feature selector (row) in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            feature_selector = row[0]
            feature_importance = np.array(row[2:])

            # Select the top 10% most influential features
            num_features = len(feature_importance)
            top_n = max(1, num_features // (1 / top_precent))  # At least one feature
            top_indices = np.argsort(np.abs(feature_importance))[-top_n:]

            # Subset the dataset with the selected features
            X_train_subset = X_train[:, top_indices]
            X_test_subset = X_test[:, top_indices]

            # Train SVM and evaluate performance
            scores = train_svm_on_selected_features(X_train_subset, y_train, X_test_subset, y_test, is_classification)

            # Add results for this feature selector as a single row
            result_sheet.append([feature_selector] + [scores[title] for title in score_titles])

    # Save the results to a new Excel file
    results_wb.save(f"svm_feature_selector_results_{top_precent}.xlsx")


if __name__ == "__main__":
    main()
