import numpy as np
import matplotlib.pyplot as plt 

import shap 
from explainers.gemfix import GEMFIX
from explainers.bishapley_kernel import Bivariate_KernelExplainer
from shapreg import removal, games, shapley
from explainers.MAPLE import MAPLE
from lime import lime_tabular

from pathlib import Path
import pandas as pd 
from openpyxl import load_workbook

from oak.input_measures import GaussianMeasure, EmpiricalMeasure
from oak.model_utils import oak_model, save_model
from oak.utils import get_model_sufficient_statistics, get_prediction_component
from oak.ortho_rbf_kernel import OrthogonalRBFKernel
#from OAK_shapley.examples.uci.functions_shap import Omega, tensorflow_to_torch, torch_to_tensorflow, numpy_to_torch
from oak.oak_kernel import OAKKernel

import gpflow

from SHOGP import SHOGP
#from synthesized_datasets import *
from synthesized_data import *
from datetime import datetime
import os

import tensorflow as tf

# List all available GPUs
gpus = tf.config.list_physical_devices('GPU')
if gpus:
    print(f"CUDA is available. GPUs detected: {len(gpus)}")
else:
    print("CUDA is not available.")

results_xsl = Path('agp_explanation_synthesized_v3.xlsx')
if not os.path.exists(results_xsl):
    # Create an empty Excel file if it doesn't exist
    pd.DataFrame().to_excel(results_xsl, index=False)

if __name__ == '__main__':
    np.random.seed(30)

    X_sample_no = 200  # number of sampels for generating explanation
    smaple_tbX = 500   # number of samples to be explained
    sample_no_gn = 1000 # number of generated synthesized instances 
    feature_no_gn = 10 # number of features for the synthesized instances

    # Example usage of one of the functions
    datasets=['Sine Log', 'Sine Cosine', 'Trigonometric Exponential', 'Exponential Hyperbolic'] #['Sine Log', 'Sine Cosine', 'Poly Sine', 'Squared Exponentials', 'Tanh Sine', \
              #'Trigonometric Exponential', 'Exponential Hyperbolic', 'XOR']
    for ds_name in datasets:
        #X, y, fn, feature_imp, ds_name = generate_data(n=sample_no_gn, d=feature_no_gn, datatype=ds)
        X, y, fn, feature_imp, g_train = generate_dataset(ds_name, sample_no_gn, feature_no_gn, 42, type="independent")
        
        selected_indices = np.random.choice(X.shape[0], size=smaple_tbX, replace=False)
        X_tbx = X[selected_indices, :]

        n,d = X.shape
    
        shogp = SHOGP(X, y, inte_order=5, inducing_points=200)
        #y_hat = shogp.OGP.predict(X)
        #ground_truth_var = np.var(y_hat)
        #var_normalized = np.var(shogp.OGP.scaler_y.transform(y_hat.reshape(-1,1)))
        #getSOBOL = shogp.OGP.get_sobol()
        #shapley_values_rescaled, shapley_values = shogp.global_shapley_value()
        
        shogp_values = shogp.local_Shapley_values(X_tbx)

        shogp_ranks = create_rank(np.array(shogp_values).squeeze())
        shogp_avg_ranks = np.mean(shogp_ranks[:,feature_imp], axis=1)
        shogp_mean_rank = np.mean(shogp_avg_ranks)

        '''
        ## SHAP
        X_bg = shap.sample(X, 100)
        explainer = shap.KernelExplainer(fn, X_bg, l1_reg=True)
        shap_values = explainer.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
        shap_ranks = create_rank(shap_values.squeeze())
        shap_avg_ranks = np.mean(shap_ranks[:,feature_imp], axis=1)
        shap_mean_rank = np.mean(shap_avg_ranks)

        ## Sampling SHAP
        sexplainer = shap.SamplingExplainer(fn, X_bg, l1_reg=True)
        sshap_values = sexplainer.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True, min_samples_per_feature=1)
        sshap_ranks = create_rank(sshap_values.squeeze())
        sshap_avg_ranks = np.mean(sshap_ranks[:,feature_imp], axis=1)
        sshap_mean_rank = np.mean(sshap_avg_ranks)


        ## Bivariate SHAP
        bishap = Bivariate_KernelExplainer(fn, X_bg)
        bishap_values = bishap.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
        bishap_ranks = create_rank(np.array(bishap_values).squeeze())
        bishap_avg_ranks = np.mean(bishap_ranks[:,feature_imp], axis=1)
        bishap_mean_rank = np.mean(bishap_avg_ranks)


        ## LIME, Unbiased SHAP, and MAPLE 
        lime_exp = lime_tabular.LimeTabularExplainer(X_bg, discretize_continuous=False, mode="regression")
        imputer = removal.MarginalExtension(X_bg, fn)
        exp_maple = MAPLE(X, y, X, y)

        ushap_values = np.empty_like(X_tbx)
        lime_values = np.empty_like(X_tbx)
        maple_values = np.empty_like(X_tbx)
        for i in range(X_tbx.shape[0]):
            x = X_tbx[i, ]
        
            ## Unbiased kernel shap 
            game = games.PredictionGame(imputer, x)
            values = shapley.ShapleyRegression(game, n_samples=X_sample_no, paired_sampling=False)
            ushap_values[i,:] = values.values.squeeze()

            ## LIME 
            exp = lime_exp.explain_instance(x, fn, num_samples = X_sample_no)
                
            for tpl in exp.as_list():
                lime_values[i, int(tpl[0])] = tpl[1]

            ## MAPLE
            mpl_exp = exp_maple.explain(x)
            maple_values[i,] = (mpl_exp['coefs'][1:]).squeeze()


        lime_ranks = create_rank(lime_values)
        lime_avg_ranks = np.mean(lime_ranks[:,feature_imp], axis=1)
        lime_mean_rank = np.mean(lime_avg_ranks)

        maple_ranks = create_rank(maple_values)
        maple_avg_ranks = np.mean(maple_ranks[:,feature_imp], axis=1)
        maple_mean_rank = np.mean(maple_avg_ranks)

        ushap_ranks = create_rank(ushap_values)
        ushap_avg_ranks = np.mean(ushap_ranks[:,feature_imp], axis=1)
        ushap_mean_rank = np.mean(ushap_avg_ranks)

        #plt.boxplot([shogp_avg_ranks, shap_avg_ranks, bishap_avg_ranks, sshap_avg_ranks, ushap_avg_ranks, lime_avg_ranks, maple_avg_ranks])

        '''
        method_names = ['SHOGP']#, 'Kernel SHAP', 'Sampling SHAP', 'Unbiased SHAP', 'Bivariate SHAP', 'LIME',  'MAPLE']
        all_results = [shogp_avg_ranks]#, shap_avg_ranks, sshap_avg_ranks, ushap_avg_ranks, bishap_avg_ranks, lime_avg_ranks, maple_avg_ranks]

        df = pd.DataFrame(all_results, index=method_names)

        mode = 'a' if results_xsl.exists() else 'w'
        # Load the existing Excel file
        book = load_workbook(results_xsl)
        
        # Remove the sheet if it already exists
        if ds_name in book.sheetnames:
            del book[ds_name]
        
        # Write the DataFrame to a new sheet
        with pd.ExcelWriter(results_xsl, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=ds_name)

    print("done!")
    

