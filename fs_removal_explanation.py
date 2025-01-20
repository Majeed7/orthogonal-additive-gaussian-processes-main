import os
import pickle
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.utils.multiclass import type_of_target
import shap
import numpy as np
from sklearn.utils.multiclass import type_of_target

from real_datasets import load_dataset
from oak.model_utils import load_model
from SHOGP import SHOGP  # Assuming SHOGP class is stored in SHOGP_class.py
from joblib import Parallel, delayed

from util import *

def reload_dataset(dataset_name):
    """Reload the dataset by its name."""
    return load_dataset(dataset_name)

def load_shogp_model(model_path):
    """Load the SHOGP model."""
    with open(model_path, 'rb') as f:
        shogp_model = pickle.load(f)
    return shogp_model

def remove_feature_and_predict(X, feature_index, shogp_model):
    """Remove a feature and make a prediction using the SHOGP model."""
    X_modified = np.delete(X, feature_index, axis=1)
    prediction = shogp_model.predict(X_modified)
    return prediction

def calculate_shap_values(fn, X_bg, X_tbx, X_sample_no=100):
    """Calculate SHAP values."""
    explainer = shap.KernelExplainer(fn, X_bg, l1_reg=True)
    shap_values = explainer.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
    shap_ranks = np.argsort(-np.abs(shap_values).mean(axis=0))  # Rank features by mean absolute SHAP values
    return shap_ranks

def calculate_sampling_shap_values(fn, X_bg, X_tbx, X_sample_no=100):
    """Calculate Sampling SHAP values."""
    sexplainer = shap.SamplingExplainer(fn, X_bg, l1_reg=True)
    sshap_values = sexplainer.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True, min_samples_per_feature=1)
    sshap_ranks = np.argsort(np.abs(sshap_values).mean(axis=0))  # Rank features by mean absolute Sampling SHAP values
    return sshap_ranks

def calculate_bivariate_shap_values(fn, X_bg, X_tbx, X_sample_no=100):
    """Calculate Bivariate SHAP values."""
    bishap = Bivariate_KernelExplainer(fn, X_bg)
    bishap_values = bishap.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
    bishap_ranks = np.argsort(np.abs(bishap_values).mean(axis=0))  # Rank features by mean absolute Bivariate SHAP values
    return bishap_ranks

def remove_features_and_predict(X, shogp_model, ranked_features, num_features_to_remove):
    """Remove the most important features and make a prediction."""
    # Remove the top `num_features_to_remove` features based on ranking
    selected_features = ranked_features[:-num_features_to_remove]
    X_modified = X[:, selected_features]  # Modify dataset by removing selected features
    prediction = shogp_model.predict(X_modified)  # Get prediction for the remaining features
    return prediction

def shogp_predict(shogp_model):
    return prediction_featuresubset(shogp_model)

def process_local_explanations_for_samples(X, y, shogp_model, explanation_method, num_samples=30, X_sample_no=100):
    """Process local explanations for 30 samples and track predictions after feature removal."""
    predictions = []
    num_features = X.shape[1]
    
    # Select 30 random samples from the dataset
    indices = np.random.choice(X.shape[0], size=num_samples, replace=False)

    # Loop through each selected sample
    for idx in indices:
        X_sample = X[idx:idx+1]
        y_sample = y[idx:idx+1]

        # Create a background set for SHAP
        X_bg = shap.sample(X, 100)  # Background dataset for SHAP

        def predict_fun(x):
            # Check if x is two-dimensional
            if x.ndim == 2:
                # Define a function to apply prediction_featuresubset for each row
                def get_prediction_for_row(row):
                    return prediction_featuresubset(shogp_model, row, np.ones(row.shape))
                
                # Parallelize the loop over the rows
                #predictions = Parallel(n_jobs=-1)(delayed(get_prediction_for_row)(x[i]) for i in range(x.shape[0]))
                predictions = []
                for i in range(x.shape[0]):
                    predictions.append(get_prediction_for_row(x[i]))
                
                # Concatenate the predictions along the first axis (rows)
                return np.concatenate(predictions, axis=0)
            else:
                # If x is one-dimensional, just run prediction once
                return prediction_featuresubset(shogp_model, x.squeeze(), np.ones(x.shape).squeeze())

        predict_fun(X[:1,:])

        # Apply the explanation method to get feature importance
        feature_ranks = calculate_shap_values(predict_fun, X_bg, X_sample, X_sample_no)

        # Track the predictions after removing features one by one
        sample_predictions = []
        for num_features_to_remove in range(1, num_features + 1):
            prediction = remove_features_and_predict(X_sample, shogp_model, feature_ranks, num_features_to_remove)
            sample_predictions.append(prediction)

        predictions.append(sample_predictions)

    return predictions

def save_predictions_to_excel(predictions, file_name="feature_removal_predictions.xlsx"):
    """Save the feature removal predictions to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Removal Predictions"

    # Column titles for prediction results (for each method)
    ws.append(["Sample Index", "SHAP Prediction", "Sampling SHAP Prediction", "Bivariate SHAP Prediction"])

    # Add predictions to the sheet
    for idx, sample_predictions in enumerate(predictions):
        for num_features_to_remove in range(len(sample_predictions['shap'])):
            ws.append([
                idx + 1,  # Sample Index
                sample_predictions['shap'][num_features_to_remove],
                sample_predictions['sampling_shap'][num_features_to_remove],
                sample_predictions['bivariate_shap'][num_features_to_remove]
            ])

    wb.save(file_name)

def main():
    # List of datasets
    datasets = ['breast_cancer', 'diabetes', 'wine_quality']  # Add more datasets as needed

    # Define explanation methods
    explanation_methods = {
        'shap': calculate_shap_values,
        'sampling_shap': calculate_sampling_shap_values,
        'bivariate_shap': calculate_bivariate_shap_values
    }

    # Create a new workbook for storing results
    wb = Workbook()

    for dataset_name in datasets:
        # Load dataset
        X, y = load_dataset(dataset_name)

        is_classification = type_of_target(y) in ["binary", "multiclass"]
        folder = 'classification' if is_classification else 'regression'
        if is_classification: continue

        # Load SHOGP model
        shogp_model = load_shogp_model(f"trained_models/{folder}/shogp_{dataset_name}.pkl")

        # Create a new sheet for the current dataset
        sheet = wb.create_sheet(title=dataset_name)
        sheet.append(["Method", "Predictions"])

        for method_name, explanation_method in explanation_methods.items():
            print(f"Processing {dataset_name} with method: {method_name}")
    
            # Process local explanations for samples
            predictions = process_local_explanations_for_samples(X, y, shogp_model, None)

            # Save the predictions to the sheet
            sheet.append([method_name] + predictions)

    # Save the workbook to an Excel file
    wb.save("feature_removal_predictions.xlsx")

if __name__ == "__main__":
    main()