import os
import pickle
import numpy as np
import pandas as pd
from sklearn.svm import SVC, SVR
from sklearn.metrics import accuracy_score, mean_squared_error, mean_absolute_percentage_error
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook, Workbook
from sklearn.preprocessing import LabelEncoder
from sklearn.utils.multiclass import type_of_target
import sys
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.model_selection import GridSearchCV
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier

from real_datasets import load_dataset


def reload_dataset(dataset_name):
    """Reload the dataset by its name."""
    return load_dataset(dataset_name)


def calculate_classification_scores(y_true, y_pred):
    """Calculate classification accuracy."""
    return {"Accuracy": accuracy_score(y_true, y_pred)}


def calculate_regression_scores(y_true, y_pred):
    """Calculate regression MSE."""
    return {"MSE": mean_absolute_percentage_error(y_true, y_pred)}


def load_best_svm_model(model_path='trained_models/best_svm_model.pkl'):
    """Load the best SVM model saved using pickle."""
    with open(model_path, 'rb') as f:
        model = pickle.load(f)
    return model


def get_svm_hyperparameters(model):
    """Extract the hyperparameters from the pre-trained SVM."""
    return model.get_params()


def train_svm_on_selected_features(X_train, y_train, X_test, y_test, svm_params, is_classification):
    """Train an SVM with the provided hyperparameters and evaluate performance."""
    # Create the SVM model with the pre-trained hyperparameters
    # Define the parameter grid for fine-tuning
    if is_classification:
        param_grid = {
            "svm__C": [0.1, 1, 10, 100],  # Regularization parameter
            "svm__gamma": [0.5, 1, 2],  # Kernel coefficient
        }
    else:
        param_grid = {
            "svm__C": [0.1, 1, 10, 100, 1000],  # Regularization parameter
            "svm__gamma": [0.5, 1, 2],  # Kernel coefficient
            'svm__epsilon': [0.1, 0.2, 0.5]
        }
        
    # Create the SVM model
    model = SVC(probability=True) if is_classification else SVR()

    # Create the pipeline
    pipeline = Pipeline([
        ("imputer", SimpleImputer(strategy="mean")),  # Handle missing values
        ("svm", model)  # SVM model
    ])

    # Perform Grid Search for hyperparameter tuning
    grid_search = GridSearchCV(pipeline, param_grid, cv=3, n_jobs=-1, scoring="accuracy" if is_classification else "neg_mean_squared_error")
    grid_search.fit(X_train, y_train)

    # Get the best estimator
    best_model = grid_search.best_estimator_

    # Predict on the test set
    y_pred = best_model.predict(X_test)

    # Calculate performance
    if is_classification:
        return calculate_classification_scores(y_test, y_pred)
    else:
        return calculate_regression_scores(y_test, y_pred)


def train_svm_on_selected_features2(X_train, y_train, X_test, y_test, svm_params, is_classification):
    if is_classification:
        model = RandomForestClassifier(random_state=42, n_jobs=-1)
    else:
        model = RandomForestRegressor(random_state=42, n_jobs=-1)

    # Define a reduced hyperparameter grid for grid search
    param_grid = {
        'model__n_estimators': [500, 1000],  # Number of trees
        'model__max_depth': [None, 10, 20],  # Maximum depth of trees
        # Minimum samples required to split an internal node
        # 'model__min_samples_split': [2, 10],
    }

    pipeline = Pipeline([
        ("imputer", SimpleImputer(strategy="mean")),  # Handle missing values
        ("model", model)
    ])

    # Set up GridSearchCV
    grid_search = GridSearchCV(estimator=pipeline, param_grid=param_grid, cv=3,
                               n_jobs=-1, verbose=2, scoring='accuracy' if is_classification else 'r2')

    # Perform grid search to find best hyperparameters
    grid_search.fit(X_train, y_train)

    best_model = grid_search.best_estimator_

    # Predict on the test set
    y_pred = best_model.predict(X_test)

    # Calculate performance
    if is_classification:
        return calculate_classification_scores(y_test, y_pred)
    else:
        return calculate_regression_scores(y_test, y_pred)


def select_features_incrementally(X_train, y_train, X_test, y_test, ranked_features, svm_params, is_classification):
    """Select features incrementally and evaluate performance."""

    performance = []
    selected_features = []

    i = 2
    while i < X_train.shape[1]:
        # Select the top `i` features
        selected_features = ranked_features[:i]

        # Subset the data for training and testing
        X_train_subset = X_train[:, selected_features]
        X_test_subset = X_test[:, selected_features]

        # Train the SVM on the selected features and evaluate
        scores = train_svm_on_selected_features(
            X_train_subset, y_train, X_test_subset, y_test, svm_params, is_classification)

        # Track performance for this number of selected features
        performance.append({'num_features': i, **scores})

        # Increment features based on the specified conditions
        if i < 25:
            i += 3
        elif i < 80:
            i += 5
        else:
            i += 10

        # Ensure i does not exceed the total number of features
        if i > X_train.shape[1]:
            i = X_train.shape[1]

    return performance


def main():
    # Load the Excel file with feature importance data
    feature_importance_file = "results/fs_real/feature_importance.xlsx"
    wb = load_workbook(feature_importance_file)

    # Create a new workbook for storing results
    results_wb = Workbook()

    # Process datasets in the Excel sheet
    for sheet_name in wb.sheetnames:
        try:
            if sheet_name in ['keggdirected']: continue 
            print(f"Processing dataset: {sheet_name}")
            sheet = wb[sheet_name]

            # Reload the dataset
            X, y = reload_dataset(sheet_name)

            # Determine if it's classification or regression
            is_classification = type_of_target(y) in ["binary", "multiclass"]
            if is_classification:
                y = LabelEncoder().fit_transform(y)
            #else:
            #    continue

            ## Load the best pre-trained SVM model to get its hyperparameters
            #pre_trained_svm = load_best_svm_model(
            #    model_path=f'trained_models/regression/svm_{sheet_name}.pkl')
            #svm_params = get_svm_hyperparameters(pre_trained_svm)
            svm_params = None

            # Train-test split
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.1, random_state=42)

            result_sheet = results_wb.create_sheet(title=sheet_name)

            # Set column titles dynamically
            score_titles = ["Accuracy"] if is_classification else ["MSE"]
            result_sheet.append(["Feature Selector"] + score_titles)

            # Process each feature selector (row) in the sheet
            for row in sheet.iter_rows(min_row=2, values_only=True):
                feature_selector = row[0]
                if feature_selector is None:
                    break
                feature_importance = np.array(row[2:])

                # Use the ranking directly from the feature_importance array
                # Sorting by importance
                ranked_features = np.argsort(-np.abs(feature_importance))

                # Incrementally evaluate performance with the selected features
                performance = select_features_incrementally(
                    X_train, y_train, X_test, y_test, ranked_features, svm_params, is_classification)

                row_index = 1
                items = ["Feature Selector"] + [result['num_features']
                                                for result in performance]
                # Insert the items in one row
                for col_index, item in enumerate(items, start=1):
                    result_sheet.cell(row=row_index, column=col_index, value=item)

                # Add results for this feature selector
                # for result in performance:
                result_sheet.append(
                    [feature_selector] + [result[score_titles[0]] for result in performance])

                # Save the results to a new Excel file
                results_wb.save(f"incremental_feature_addition.xlsx")

        except:
            print(f"{sheet_name} could not be processed!")
            continue

if __name__ == "__main__":
    main()
