from sklearn.feature_selection import mutual_info_classif, RFECV, SelectKBest, chi2, mutual_info_regression
from synthesized_datasets import *
from sklearn.linear_model import Lasso
from sklearn.feature_selection import SelectFromModel
from sklearn.svm import SVC, SVR
from sklearn.model_selection import StratifiedKFold
from sklearn.ensemble import ExtraTreesClassifier, ExtraTreesRegressor
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.feature_selection import SelectKBest, f_classif, f_regression
import numpy as np
from pyHSICLasso import HSICLasso
from pathlib import Path
from datetime import datetime
import os 
from openpyxl import load_workbook
import time

from SHOGP import SHOGP
from synthesized_data import *

import tensorflow as tf

# List all available GPUs
gpus = tf.config.list_physical_devices('GPU')
if gpus:
    print(f"CUDA is available. GPUs detected: {len(gpus)}")
else:
    print("CUDA is not available.")

results_xsl = Path('fs_synthesized.xlsx')
if not os.path.exists(results_xsl):
    # Create an empty Excel file if it doesn't exist
    pd.DataFrame().to_excel(results_xsl, index=False)



if __name__ == '__main__':
    np.random.seed(30)

    X_sample_no = 500  # number of sampels for generating explanation
    smaple_tbX = 200   # number of samples to be explained
    sample_no_gn = 2000 # number of generated synthesized instances 
    feature_no_gn = 20 # number of features for the synthesized instances

    exp_no = 2 
    importance_mi = np.zeros((exp_no,feature_no_gn))
    importance_lasso = np.zeros((exp_no,feature_no_gn))
    orders_rfecv = np.zeros((exp_no,feature_no_gn))
    importance_k_best = np.zeros((exp_no,feature_no_gn))
    importance_ensemble = np.zeros((exp_no,feature_no_gn))
    importance_shogp = np.zeros((exp_no,feature_no_gn))
    importance_sobol = np.zeros((exp_no,feature_no_gn))
    importance_hsiclasso = np.zeros((exp_no,feature_no_gn))

    time_mi = np.zeros(exp_no)
    time_lasso = np.zeros(exp_no)
    time_rfecv = np.zeros(exp_no)
    time_k_best = np.zeros(exp_no)
    time_ensemble = np.zeros(exp_no)
    time_shogp = np.zeros(exp_no)
    time_shogp_train = np.zeros(exp_no)
    time_sobol = np.zeros(exp_no)
    time_hsiclasso = np.zeros(exp_no)

    # Example usage of one of the functions
    datasets=['Sine Log', 'Sine Cosine', 'Poly Sine', 'Squared Exponentials', 'Tanh Sine', 
              'Trigonometric Exponential', 'Exponential Hyperbolic', 'XOR']
    
    for ds_name in datasets:
            for i in range(exp_no):
                X, y, fn, feature_imp, g_train = generate_dataset(ds_name, sample_no_gn, feature_no_gn, 42)

                mode = 'regression'

                ## SHOGP Shpalley values and sobol indices
                start_time = time.time()
                shogp = SHOGP(X, y, inte_order=5)
                time_shogp_train[i] = time.time() - start_time  

                start_time = time.time()
                shapley_values_rescaled, shapley_values = shogp.global_shapley_value()
                time_shogp[i] = time.time() - start_time 
                importance_shogp[i,:] = shapley_values

                start_time = time.time()
                importance_sobol[i,:] = shogp.get_sobol()
                time_sobol[i] = time.time() - start_time 

                ## HSIC lasso 
                start_time = time.time()
                hsic_lasso = HSICLasso()
                hsic_lasso.input(X,y)
                hsic_lasso.regression(feature_no_gn, covars=X)
                hsic_ind = hsic_lasso.get_index()
                init_ranks = (len(hsic_ind) + (feature_no_gn - 1/2 - len(hsic_ind))/2) * np.ones((feature_no_gn,))
                init_ranks[hsic_ind] = np.arange(1,len(hsic_ind)+1)
                importance_hsiclasso[i,:] = init_ranks 
                time_hsiclasso[i] = time.time() - start_time 

                ## Mutual Informmation Importance
                start_time = time.time()
                importance_mi[i,:] = mutual_info_classif(X,y) if mode == 'classification' else mutual_info_regression(X,y)
                time_mi[i] = time.time() - start_time

                ## Lasso importance
                start_time = time.time()
                lasso = Lasso().fit(X, y)
                importance_lasso[i,:] = np.abs(lasso.coef_)
                time_lasso[i] = time.time() - start_time

                #Recursive elimination
                start_time = time.time()
                estimator = SVC(kernel="linear") if mode == 'classification' else SVR(kernel='linear')
                rfecv = RFECV(estimator, step=1, cv=5)
                rfecv.fit(X, y)
                orders_rfecv[i,:] = rfecv.ranking_
                time_rfecv[i] = time.time() - start_time    

                ## K best
                start_time = time.time()
                bestfeatures = SelectKBest(score_func=f_classif, k='all') if mode == 'classification' else SelectKBest(score_func=f_regression, k='all') #F-ANOVA feature selection
                fit = bestfeatures.fit(X,y)
                importance_k_best[i,:] = fit.scores_
                time_k_best[i] = time.time() - start_time   

                ## Tree ensemble 
                start_time = time.time()
                model = ExtraTreesClassifier(n_estimators=50) if mode =='classification' else ExtraTreesRegressor(n_estimators=50)
                model.fit(X,y)
                importance_ensemble[i,:] = model.feature_importances_
                time_ensemble[i] = time.time() - start_time 

            ranking_mi = create_rank(importance_mi)
            ranking_lasso = create_rank(importance_lasso)
            ranking_k_best = create_rank(importance_k_best)
            ranking_ensemble = create_rank(importance_ensemble)
            ranking_rfecv = create_rank(orders_rfecv)
            ranking_hsiclasso = importance_hsiclasso
            ranking_shogp = create_rank(np.abs(importance_shogp))
            ranking_sobol = create_rank(np.abs(importance_sobol))
            
            avg_mi = (np.mean(ranking_mi[:,feature_imp],axis=1))
            avg_lasso = (np.mean(ranking_lasso[:,feature_imp],axis=1))
            avg_k_best = (np.mean(ranking_k_best[:,feature_imp],axis=1))
            avg_ensemble = (np.mean(ranking_ensemble[:,feature_imp],axis=1))
            avg_rfecv = (np.mean(ranking_rfecv[:,feature_imp],axis=1))
            avg_hsic_lasso = (np.mean(ranking_hsiclasso[:,feature_imp],axis=1))
            avg_shogp = (np.mean(ranking_shogp[:,feature_imp],axis=1))
            avg_sobol = (np.mean(ranking_sobol[:,feature_imp],axis=1))  
        
            # Creating dataset
            data = [avg_shogp, time_shogp, time_shogp_train, avg_sobol, time_sobol, avg_hsic_lasso, time_hsiclasso, avg_mi, time_mi, avg_k_best, time_k_best, \
                     avg_rfecv, time_rfecv, avg_lasso, time_lasso, avg_ensemble, time_ensemble]
            methods = ["SHOGP", "SHOGP_time", "SHOGP_time_train", "Sobol", "Sobol_time", "HSIC-Lasso", "HSIC_time", "MI","MI_time", "F-ANOVA", "F_ANOVA_time", \
                       "RFEC", "RFEC_time", "Lasso", "Lasso_time", "Tree Ensemble", "Tree_ensemble_time"]

            df = pd.DataFrame(data, index=methods)

            mode = 'a' if results_xsl.exists() else 'w'
            # Load the existing Excel file
            book = load_workbook(results_xsl)
            
            # Remove the sheet if it already exists
            if ds_name in book.sheetnames:
                del book[ds_name]
            
            # Write the DataFrame to a new sheet
            with pd.ExcelWriter(results_xsl, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, index=False, sheet_name=ds_name)


# print('done')

